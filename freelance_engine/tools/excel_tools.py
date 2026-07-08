#!/usr/bin/env python3
"""
Excel 自动化工具
用途：Excel 数据清洗/合并/拆分/报表
接单价格：¥100-300/单
"""
import sys, os, json, re
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def merge_excel(files, output):
    """合并多个 Excel 文件"""
    all_dfs = []
    for f in files:
        df = pd.read_excel(f)
        all_dfs.append(df)
    result = pd.concat(all_dfs, ignore_index=True)
    result.to_excel(output, index=False)
    return {"success": True, "rows": len(result), "file": output}

def split_excel(file, column, output_dir):
    """按列值拆分 Excel"""
    df = pd.read_excel(file)
    os.makedirs(output_dir, exist_ok=True)
    result_files = []
    for val, group in df.groupby(column):
        out = os.path.join(output_dir, f"{val}.xlsx")
        group.to_excel(out, index=False)
        result_files.append(out)
    return {"success": True, "files": result_files, "count": len(result_files)}

def clean_excel(file, output):
    """数据清洗：去重、补空、统一格式"""
    df = pd.read_excel(file)
    before = len(df)
    df = df.drop_duplicates()
    df = df.fillna("")
    # 自动识别数值列并修复格式
    for col in df.columns:
        if df[col].dtype == object:
            df[col] = df[col].astype(str).str.strip()
    df.to_excel(output, index=False)
    stats = {"总行数": before, "去重后": len(df), "删除": before - len(df)}
    return {"success": True, "stats": stats, "file": output}

def analyze_excel(file, output):
    """数据分析 + 生成统计报告 Excel"""
    df = pd.read_excel(file)
    wb = Workbook()
    # Sheet 1: 原始数据
    ws1 = wb.active
    ws1.title = "原始数据"
    for r_idx, row in enumerate(df.values, 1):
        for c_idx, val in enumerate(row, 1):
            ws1.cell(r_idx, c_idx, val)
    # Sheet 2: 统计信息
    ws2 = wb.create_sheet("统计信息")
    ws2.append(["字段", "数据类型", "非空数", "唯一值", "均值(数值)", "最大值", "最小值"])
    for col in df.columns:
        vals = df[col]
        ws2.append([
            col, str(df[col].dtype), int(vals.count()),
            vals.nunique(),
            round(vals.mean(), 2) if vals.dtype in ["float64", "int64"] else "",
            vals.max() if vals.dtype in ["float64", "int64"] else "",
            vals.min() if vals.dtype in ["float64", "int64"] else ""
        ])
    # 美化
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F6EF7")
    for cell in ws2[1]:
        cell.font = header_font
        cell.fill = header_fill
    ws2.column_dimensions["A"].width = 20
    wb.save(output)
    return {"success": True, "columns": len(df.columns), "rows": len(df), "file": output}

def format_excel(file, output):
    """美化 Excel 样式"""
    wb = load_workbook(file)
    header_fill = PatternFill("solid", fgColor="4F6EF7")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.border = thin_border
        # 自动列宽
        for col_idx in range(1, ws.max_column + 1):
            max_len = 0
            col_letter = get_column_letter(col_idx)
            for cell in ws[col_letter]:
                try:
                    max_len = max(max_len, len(str(cell.value or "")))
                except:
                    pass
            ws.column_dimensions[col_letter].width = min(max_len + 2, 50)
    wb.save(output)
    return {"success": True, "file": output}

def main():
    if len(sys.argv) < 3:
        print(json.dumps({"error": "用法: excel_tools.py <命令> <参数...>", "命令": ["merge", "split", "clean", "analyze", "format"]}))
        return
    cmd = sys.argv[1]
    try:
        if cmd == "merge":
            files = sys.argv[2].split("|")
            out = sys.argv[3] if len(sys.argv) > 3 else "merged.xlsx"
            result = merge_excel(files, out)
        elif cmd == "split":
            result = split_excel(sys.argv[2], sys.argv[3], sys.argv[4] if len(sys.argv) > 4 else "output")
        elif cmd == "clean":
            result = clean_excel(sys.argv[2], sys.argv[3] if len(sys.argv) > 3 else "cleaned.xlsx")
        elif cmd == "analyze":
            result = analyze_excel(sys.argv[2], sys.argv[3] if len(sys.argv) > 3 else "analysis.xlsx")
        elif cmd == "format":
            result = format_excel(sys.argv[2], sys.argv[3] if len(sys.argv) > 3 else "formatted.xlsx")
        else:
            result = {"error": f"未知命令: {cmd}"}
    except Exception as e:
        result = {"error": str(e)}
    print(json.dumps(result, ensure_ascii=False))

if __name__ == "__main__":
    main()
